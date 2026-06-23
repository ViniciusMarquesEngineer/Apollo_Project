"""
Apollo Logística — Servidor Flask
Execute: python app.py
Acesse:  http://localhost:5000
"""

from flask import Flask, request, jsonify, send_from_directory, session
from werkzeug.security import generate_password_hash, check_password_hash
import sqlite3, os, io, random, string, json, base64
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ── QR Code: geração via qrcode + pyzbar, leitura de câmera via cv2 ──────────
import threading

try:
    import qrcode as _qrcode_lib
    QR_DISPONIVEL = True
except ImportError:
    QR_DISPONIVEL = False
    print("[Apollo] Módulo 'qrcode' não encontrado.")

# Leitura de QR feita no browser via jsQR (sem dependência de OpenCV)

# ─────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH  = os.path.join(BASE_DIR, 'apollo.db')

app = Flask(__name__, static_folder='static', template_folder='templates')
app.secret_key = os.environ.get('SECRET_KEY', 'apollo_secret_2026_piec')

# ── Configuração de sessão/cookie para funcionar em HTTPS (Render/produção) ──
is_production = os.environ.get('RENDER', '') != ''   # Render define a var RENDER
app.config.update(
    SESSION_COOKIE_SECURE    = is_production,   # só envia cookie por HTTPS
    SESSION_COOKIE_HTTPONLY  = True,            # bloqueia acesso JS ao cookie
    SESSION_COOKIE_SAMESITE  = 'Lax',          # compatível com redirect pós-login
    PERMANENT_SESSION_LIFETIME = 86400,         # sessão dura 24h
)

# ═══════════════════════════════════════════════════════════
#  BANCO DE DADOS
# ═══════════════════════════════════════════════════════════
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    c = conn.cursor()

    # Usuários
    c.execute('''CREATE TABLE IF NOT EXISTS usuarios (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        usuario     TEXT UNIQUE NOT NULL,
        email       TEXT UNIQUE NOT NULL,
        senha_hash  TEXT NOT NULL,
        role        TEXT NOT NULL DEFAULT 'funcionario',
        verificado  INTEGER NOT NULL DEFAULT 0,
        criado_em   TEXT NOT NULL
    )''')

    # Códigos de verificação de e-mail
    c.execute('''CREATE TABLE IF NOT EXISTS codigos_verificacao (
        id        INTEGER PRIMARY KEY AUTOINCREMENT,
        email     TEXT NOT NULL,
        codigo    TEXT NOT NULL,
        expira_em TEXT NOT NULL,
        usado     INTEGER NOT NULL DEFAULT 0
    )''')

    # Produtos / estoque
    c.execute('''CREATE TABLE IF NOT EXISTS estoque (
        codigo                TEXT PRIMARY KEY,
        nome                  TEXT NOT NULL,
        categoria             TEXT NOT NULL,
        quantidade_atual      INTEGER NOT NULL DEFAULT 0,
        estoque_minimo        INTEGER NOT NULL DEFAULT 0,
        fornecedor            TEXT DEFAULT '',
        fornecedor_email      TEXT DEFAULT '',
        fornecedor_telefone   TEXT DEFAULT '',
        atualizado_em         TEXT
    )''')
    # Migração: garante coluna mesmo em bancos antigos
    try:
        c.execute("ALTER TABLE estoque ADD COLUMN fornecedor_telefone TEXT DEFAULT ''")
    except Exception:
        pass  # coluna já existe

    # Movimentações
    c.execute('''CREATE TABLE IF NOT EXISTS movimentacoes (
        id              INTEGER PRIMARY KEY AUTOINCREMENT,
        codigo_produto  TEXT NOT NULL,
        nome_produto    TEXT NOT NULL,
        categoria       TEXT NOT NULL,
        tipo            TEXT NOT NULL,
        quantidade      INTEGER NOT NULL,
        estoque_apos    INTEGER NOT NULL,
        data            TEXT NOT NULL,
        operador        TEXT DEFAULT '',
        registrado_em   TEXT NOT NULL,
        fornecedor      TEXT DEFAULT '',
        FOREIGN KEY(codigo_produto) REFERENCES estoque(codigo)
    )''')
    # Migração: adiciona coluna fornecedor se não existir (bancos antigos)
    try:
        c.execute("ALTER TABLE movimentacoes ADD COLUMN fornecedor TEXT DEFAULT ''")
    except Exception:
        pass

    conn.commit()
    conn.close()
    print("[Apollo] Banco de dados iniciado em:", DB_PATH)

# ═══════════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════════
def usuario_logado():
    return session.get('usuario_id') is not None

def somente_admin():
    return session.get('role') == 'admin'

def gerar_codigo_verificacao():
    return ''.join(random.choices(string.digits, k=6))

def resp_erro(msg, code=400):
    return jsonify({'ok': False, 'erro': msg}), code

def resp_ok(dados=None, msg=''):
    r = {'ok': True}
    if msg:   r['msg']   = msg
    if dados: r['dados'] = dados
    return jsonify(r)

# ═══════════════════════════════════════════════════════════
#  SERVIR PÁGINAS HTML
# ═══════════════════════════════════════════════════════════
@app.route('/')
def index():
    return send_from_directory('templates', 'apollo-landing.html')

@app.route('/dashboard')
def dashboard():
    return send_from_directory('templates', 'apollo-dashboard.html')

@app.route('/admin')
def admin():
    return send_from_directory('templates', 'apollo-admin.html')

@app.route('/static/<path:filename>')
def static_files(filename):
    return send_from_directory('static', filename)

# ═══════════════════════════════════════════════════════════
#  AUTH — CADASTRO
# ═══════════════════════════════════════════════════════════
@app.route('/api/auth/cadastro/solicitar', methods=['POST'])
def cadastro_solicitar():
    """
    Etapa 1: valida dados e gera código de verificação.
    SEGURANÇA: o role enviado pelo cliente é completamente ignorado.
    Todo cadastro público é sempre 'funcionario'.
    Apenas um admin autenticado pode promover usuários via /api/usuarios/role.
    """
    d = request.get_json()
    usuario = (d.get('usuario') or '').strip()
    email   = (d.get('email')   or '').strip().lower()
    senha   = (d.get('senha')   or '')
    conf    = (d.get('confirmar_senha') or '')

    # Role SEMPRE forçado para funcionario — nunca vem do cliente
    role = 'funcionario'

    if not all([usuario, email, senha, conf]):
        return resp_erro('Preencha todos os campos.')
    if senha != conf:
        return resp_erro('As senhas não coincidem.')
    import re as _re
    if len(senha) < 8:
        return resp_erro('A senha deve ter pelo menos 8 caracteres.')
    if not _re.search(r'[A-Z]', senha):
        return resp_erro('A senha deve conter pelo menos 1 letra maiuscula.')
    if not _re.search(r'[!@#$%^&*()\-_=+\[\]{};:\'",.<>/?\\|`~]', senha):
        return resp_erro('A senha deve conter pelo menos 1 caractere especial (ex: !@#$%).')

    conn = get_db()
    c    = conn.cursor()
    if c.execute('SELECT id FROM usuarios WHERE email=? OR usuario=?', (email, usuario)).fetchone():
        conn.close()
        return resp_erro('E-mail ou usuário já cadastrado.')

    # Gera código e salva (expira em 10 min)
    codigo    = gerar_codigo_verificacao()
    expira_em = (datetime.now() + timedelta(minutes=10)).isoformat()
    c.execute('DELETE FROM codigos_verificacao WHERE email=?', (email,))
    c.execute('INSERT INTO codigos_verificacao (email, codigo, expira_em) VALUES (?,?,?)',
              (email, codigo, expira_em))
    conn.commit()
    conn.close()

    # Guarda dados na sessão — role já é funcionario, não vem do body
    session['pendente'] = {'usuario': usuario, 'email': email,
                           'senha': senha, 'role': role}

    print(f"[Apollo] Código de verificação para {email}: {codigo}")
    return resp_ok({'codigo_debug': codigo}, 'Código gerado. Verifique seu e-mail.')


@app.route('/api/auth/cadastro/verificar', methods=['POST'])
def cadastro_verificar():
    """Etapa 2: verifica código e cria a conta"""
    d      = request.get_json()
    codigo = (d.get('codigo') or '').strip()
    pend   = session.get('pendente')

    if not pend:
        return resp_erro('Sessão expirada. Reinicie o cadastro.')
    if not codigo or len(codigo) != 6:
        return resp_erro('Digite os 6 dígitos do código.')

    email = pend['email']
    conn  = get_db()
    c     = conn.cursor()

    row = c.execute(
        'SELECT * FROM codigos_verificacao WHERE email=? AND usado=0 ORDER BY id DESC LIMIT 1',
        (email,)
    ).fetchone()

    if not row:
        conn.close()
        return resp_erro('Código não encontrado. Solicite um novo.')
    if datetime.fromisoformat(row['expira_em']) < datetime.now():
        conn.close()
        return resp_erro('Código expirado. Solicite um novo.')
    if row['codigo'] != codigo:
        conn.close()
        return resp_erro('Código incorreto.')

    # Marca código como usado
    c.execute('UPDATE codigos_verificacao SET usado=1 WHERE id=?', (row['id'],))

    # Cria usuário
    senha_hash = generate_password_hash(pend['senha'])
    criado_em  = datetime.now().isoformat()
    c.execute(
        'INSERT INTO usuarios (usuario, email, senha_hash, role, verificado, criado_em) VALUES (?,?,?,?,1,?)',
        (pend['usuario'], pend['email'], senha_hash, pend['role'], criado_em)
    )
    conn.commit()
    user_id = c.lastrowid
    conn.close()

    # Limpa pendente e inicia sessão
    session.pop('pendente', None)
    session['usuario_id'] = user_id
    session['usuario']    = pend['usuario']
    session['email']      = pend['email']
    session['role']       = pend['role']

    return resp_ok({'role': pend['role'], 'usuario': pend['usuario']}, 'Conta criada com sucesso!')


@app.route('/api/auth/cadastro/reenviar', methods=['POST'])
def cadastro_reenviar():
    """Reenvia (gera novo) código de verificação"""
    pend = session.get('pendente')
    if not pend:
        return resp_erro('Sessão expirada. Reinicie o cadastro.')

    email     = pend['email']
    codigo    = gerar_codigo_verificacao()
    expira_em = (datetime.now() + timedelta(minutes=10)).isoformat()

    conn = get_db()
    c    = conn.cursor()
    c.execute('DELETE FROM codigos_verificacao WHERE email=?', (email,))
    c.execute('INSERT INTO codigos_verificacao (email, codigo, expira_em) VALUES (?,?,?)',
              (email, codigo, expira_em))
    conn.commit()
    conn.close()

    print(f"[Apollo] Novo código para {email}: {codigo}")
    return resp_ok({'codigo_debug': codigo}, 'Novo código gerado.')


# ═══════════════════════════════════════════════════════════
#  AUTH — LOGIN / LOGOUT / SESSÃO
# ═══════════════════════════════════════════════════════════
@app.route('/api/auth/login', methods=['POST'])
def login():
    d     = request.get_json()
    ident = (d.get('identificador') or '').strip().lower()
    senha = (d.get('senha') or '')

    if not ident or not senha:
        return resp_erro('Preencha todos os campos.')

    conn = get_db()
    row  = conn.execute(
        'SELECT * FROM usuarios WHERE (email=? OR usuario=?) AND verificado=1',
        (ident, ident)
    ).fetchone()
    conn.close()

    if not row or not check_password_hash(row['senha_hash'], senha):
        return resp_erro('Credenciais inválidas.')

    session['usuario_id'] = row['id']
    session['usuario']    = row['usuario']
    session['email']      = row['email']
    session['role']       = row['role']

    return resp_ok({'role': row['role'], 'usuario': row['usuario']}, 'Login realizado!')


@app.route('/api/auth/logout', methods=['POST'])
def logout():
    session.clear()
    return resp_ok(msg='Logout realizado.')


@app.route('/api/auth/sessao', methods=['GET'])
def sessao():
    if not usuario_logado():
        return resp_erro('Não autenticado.', 401)
    return resp_ok({
        'usuario': session['usuario'],
        'email':   session['email'],
        'role':    session['role']
    })


# ═══════════════════════════════════════════════════════════
#  ESTOQUE — PRODUTOS
# ═══════════════════════════════════════════════════════════
@app.route('/api/estoque', methods=['GET'])
def listar_estoque():
    if not usuario_logado():
        return resp_erro('Não autenticado.', 401)
    conn  = get_db()
    rows  = conn.execute('SELECT * FROM estoque ORDER BY nome').fetchall()
    conn.close()
    return resp_ok([dict(r) for r in rows])


@app.route('/api/estoque/<codigo>', methods=['GET'])
def buscar_produto(codigo):
    if not usuario_logado():
        return resp_erro('Não autenticado.', 401)
    conn = get_db()
    row  = conn.execute('SELECT * FROM estoque WHERE codigo=?', (codigo,)).fetchone()
    conn.close()
    if not row:
        return resp_erro('Produto não encontrado.', 404)
    return resp_ok(dict(row))


@app.route('/api/estoque/fornecedor', methods=['PUT'])
def atualizar_fornecedor():
    if not usuario_logado() or not somente_admin():
        return resp_erro('Acesso negado.', 403)
    d      = request.get_json()
    codigo = d.get('codigo', '').strip()
    nome   = d.get('fornecedor', '').strip()
    email  = d.get('fornecedor_email', '').strip()
    tel    = d.get('fornecedor_telefone', '').strip()
    if not codigo:
        return resp_erro('Código do produto obrigatório.')
    conn = get_db()
    conn.execute(
        'UPDATE estoque SET fornecedor=?, fornecedor_email=?, fornecedor_telefone=?, atualizado_em=? WHERE codigo=?',
        (nome, email, tel, datetime.now().isoformat(), codigo)
    )
    conn.commit()
    conn.close()
    return resp_ok(msg='Fornecedor atualizado.')


# ═══════════════════════════════════════════════════════════
#  MOVIMENTAÇÕES
# ═══════════════════════════════════════════════════════════
@app.route('/api/movimentacao', methods=['POST'])
def registrar_movimentacao():
    if not usuario_logado():
        return resp_erro('Não autenticado.', 401)

    d      = request.get_json()
    codigo = (d.get('codigo') or '').strip().upper()
    nome   = (d.get('nome')   or '').strip()
    cat    = (d.get('categoria') or '').strip()
    tipo   = (d.get('tipo')   or '').strip().lower()
    qtd    = int(d.get('quantidade') or 0)
    data   = (d.get('data')   or datetime.now().strftime('%Y-%m-%d'))
    f_nome = (d.get('fornecedor') or '').strip()
    f_email= (d.get('fornecedor_email') or '').strip()
    f_tel  = (d.get('fornecedor_telefone') or '').strip()

    # estoque_minimo: só usa o valor enviado se ele foi explicitamente informado (> 0 ou campo preenchido)
    # Quando vazio na saída, preserva o valor já salvo no banco
    minimo_enviado = d.get('estoque_minimo')

    if not all([codigo, nome, cat, tipo]) or qtd < 1:
        return resp_erro('Preencha todos os campos obrigatórios.')
    if tipo not in ('entrada', 'saida'):
        return resp_erro('Tipo inválido.')

    conn = get_db()
    c    = conn.cursor()

    # Busca produto existente
    prod = c.execute('SELECT * FROM estoque WHERE codigo=?', (codigo,)).fetchone()

    if prod:
        atual = prod['quantidade_atual']
        # Preserva estoque_minimo do banco se não foi enviado ou veio vazio
        if minimo_enviado is None or str(minimo_enviado).strip() == '':
            minimo = prod['estoque_minimo']
        else:
            minimo = int(minimo_enviado)
    else:
        atual  = 0
        minimo = int(minimo_enviado) if minimo_enviado not in (None, '') else 0
        c.execute(
            'INSERT INTO estoque (codigo, nome, categoria, quantidade_atual, estoque_minimo, fornecedor, fornecedor_email, fornecedor_telefone, atualizado_em) VALUES (?,?,?,0,?,?,?,?,?)',
            (codigo, nome, cat, minimo, f_nome, f_email, f_tel, datetime.now().isoformat())
        )

    # Calcula novo estoque
    if tipo == 'entrada':
        novo_estoque = atual + qtd
    else:
        if atual < qtd:
            conn.close()
            return resp_erro(f'Estoque insuficiente! Disponível: {atual}')
        novo_estoque = atual - qtd

    # Atualiza estoque preservando campos não enviados
    upd_forn  = f_nome  if f_nome  else (prod['fornecedor']            if prod else '')
    upd_femail= f_email if f_email else (prod['fornecedor_email']      if prod else '')
    upd_ftel  = f_tel   if f_tel   else (prod['fornecedor_telefone']   if prod else '')
    c.execute(
        'UPDATE estoque SET quantidade_atual=?, estoque_minimo=?, nome=?, categoria=?, fornecedor=?, fornecedor_email=?, fornecedor_telefone=?, atualizado_em=? WHERE codigo=?',
        (novo_estoque, minimo, nome, cat, upd_forn, upd_femail, upd_ftel, datetime.now().isoformat(), codigo)
    )

    # Registra movimentação
    c.execute(
        'INSERT INTO movimentacoes (codigo_produto, nome_produto, categoria, tipo, quantidade, estoque_apos, data, operador, registrado_em, fornecedor) VALUES (?,?,?,?,?,?,?,?,?,?)',
        (codigo, nome, cat, tipo, qtd, novo_estoque, data, session.get('usuario',''), datetime.now().isoformat(), upd_forn)
    )

    # Gera QR Code em base64
    # Busca dados do fornecedor para incluir no QR
    prod_row = c.execute('SELECT fornecedor, fornecedor_email, fornecedor_telefone FROM estoque WHERE codigo=?', (codigo,)).fetchone()
    forn      = prod_row['fornecedor']           if prod_row else ''
    forn_em   = prod_row['fornecedor_email']     if prod_row else ''
    forn_tel  = prod_row['fornecedor_telefone']  if prod_row else ''
    qr_b64 = gerar_qr_base64(codigo, nome, cat, tipo, qtd, novo_estoque, minimo, data, forn, forn_em, forn_tel)

    conn.commit()
    conn.close()

    return resp_ok({
        'estoque_atual': novo_estoque,
        'qr_base64': qr_b64,
        'filename': f'qr_{codigo}_{tipo}_{data}.png'
    }, f'{"Entrada" if tipo=="entrada" else "Saída"} registrada com sucesso!')


@app.route('/api/movimentacoes', methods=['GET'])
def listar_movimentacoes():
    if not usuario_logado():
        return resp_erro('Não autenticado.', 401)

    data_de  = request.args.get('de')
    data_ate = request.args.get('ate')
    tipo     = request.args.get('tipo')
    cat      = request.args.get('categoria')

    query  = '''SELECT m.*, COALESCE(NULLIF(m.fornecedor,''), e.fornecedor, '') AS fornecedor
               FROM movimentacoes m
               LEFT JOIN estoque e ON m.codigo_produto = e.codigo
               WHERE 1=1'''
    params = []

    if data_de:
        query += ' AND data >= ?'; params.append(data_de)
    if data_ate:
        query += ' AND data <= ?'; params.append(data_ate)
    if tipo:
        query += ' AND tipo = ?'; params.append(tipo)
    if cat:
        query += ' AND categoria = ?'; params.append(cat)

    query += ' ORDER BY id DESC'

    conn = get_db()
    rows = conn.execute(query, params).fetchall()
    conn.close()
    return resp_ok([dict(r) for r in rows])


@app.route('/api/movimentacoes/hoje', methods=['GET'])
def movimentacoes_hoje():
    if not usuario_logado():
        return resp_erro('Não autenticado.', 401)
    hoje = datetime.now().strftime('%Y-%m-%d')
    conn = get_db()
    rows = conn.execute(
        '''SELECT m.*, COALESCE(NULLIF(m.fornecedor,''), e.fornecedor, '') AS fornecedor
           FROM movimentacoes m
           LEFT JOIN estoque e ON m.codigo_produto = e.codigo
           WHERE m.data=? ORDER BY m.id DESC''', (hoje,)
    ).fetchall()
    conn.close()
    return resp_ok([dict(r) for r in rows])


# ═══════════════════════════════════════════════════════════
#  ALERTAS
# ═══════════════════════════════════════════════════════════
@app.route('/api/alertas', methods=['GET'])
def alertas():
    if not usuario_logado():
        return resp_erro('Não autenticado.', 401)
    conn = get_db()
    rows = conn.execute(
        'SELECT * FROM estoque WHERE quantidade_atual <= estoque_minimo ORDER BY nome'
    ).fetchall()
    conn.close()
    return resp_ok([dict(r) for r in rows])


# ═══════════════════════════════════════════════════════════
#  USUÁRIOS (somente admin)
# ═══════════════════════════════════════════════════════════
@app.route('/api/usuarios', methods=['GET'])
def listar_usuarios():
    if not usuario_logado() or not somente_admin():
        return resp_erro('Acesso negado.', 403)
    conn = get_db()
    rows = conn.execute(
        'SELECT id, usuario, email, role, verificado, criado_em FROM usuarios ORDER BY criado_em DESC'
    ).fetchall()
    conn.close()
    return resp_ok([dict(r) for r in rows])


@app.route('/api/usuarios/role', methods=['PUT'])
def alterar_role():
    """
    Permite que um admin promova ou rebaixe outro usuário.
    SEGURANÇA:
    - Exige sessão de admin autenticado no servidor
    - Admin não pode alterar o próprio role (evita acidente)
    - Aceita apenas 'funcionario' ou 'admin' como valores válidos
    - Registra no terminal quem fez a alteração e para quem
    """
    if not usuario_logado() or not somente_admin():
        return resp_erro('Acesso negado. Somente administradores.', 403)

    d          = request.get_json()
    usuario_id = d.get('usuario_id')
    novo_role  = (d.get('role') or '').strip().lower()

    if not usuario_id:
        return resp_erro('ID de usuário obrigatório.')
    if novo_role not in ('funcionario', 'admin'):
        return resp_erro('Role inválido. Use "funcionario" ou "admin".')
    if usuario_id == session.get('usuario_id'):
        return resp_erro('Você não pode alterar o seu próprio nível de acesso.')

    conn = get_db()
    c    = conn.cursor()
    alvo = c.execute('SELECT id, usuario, role FROM usuarios WHERE id=?', (usuario_id,)).fetchone()
    if not alvo:
        conn.close()
        return resp_erro('Usuário não encontrado.')

    c.execute('UPDATE usuarios SET role=? WHERE id=?', (novo_role, usuario_id))
    conn.commit()
    conn.close()

    print(f"[Apollo] ROLE ALTERADO: admin '{session['usuario']}' alterou "
          f"'{alvo['usuario']}' de '{alvo['role']}' para '{novo_role}'")

    return resp_ok(
        {'usuario': alvo['usuario'], 'role_anterior': alvo['role'], 'role_novo': novo_role},
        f"Perfil de '{alvo['usuario']}' alterado para {novo_role} com sucesso."
    )


# ═══════════════════════════════════════════════════════════
#  GERAÇÃO DE QR CODE (qrcode + Pillow)
# ═══════════════════════════════════════════════════════════
def gerar_qr_base64(codigo, nome, cat, tipo, qtd, atual, minimo, data,
                    fornecedor='', fornecedor_email='', fornecedor_telefone=''):
    """
    Gera QR Code PNG em base64 usando a biblioteca qrcode[pil].
    Inclui dados do produto e fornecedor para preenchimento automático via câmera.
    Retorna string base64 da imagem PNG ou '' em caso de falha.
    """
    try:
        dt = datetime.strptime(data, '%Y-%m-%d').strftime('%d/%m/%Y')
    except Exception:
        dt = data

    linhas = [
        "APOLLO LOGISTICA",
        f"Codigo: {codigo}",
        f"Produto: {nome}",
        f"Categoria: {cat}",
        f"Estoque minimo: {minimo}",
    ]
    if fornecedor:
        linhas.append(f"Fornecedor: {fornecedor}")
    if fornecedor_email:
        linhas.append(f"Fornecedor Email: {fornecedor_email}")
    if fornecedor_telefone:
        linhas.append(f"Fornecedor Telefone: {fornecedor_telefone}")
    linhas.append(f"Data: {dt}")

    txt = "\n".join(linhas)

    if not QR_DISPONIVEL:
        print('[Apollo] qrcode nao instalado. Execute: pip install qrcode[pil]')
        return ''

    try:
        qr = _qrcode_lib.QRCode(
            version=None,
            error_correction=_qrcode_lib.constants.ERROR_CORRECT_M,
            box_size=8, border=4
        )
        qr.add_data(txt)
        qr.make(fit=True)
        img = qr.make_image(fill_color='#0a1f44', back_color='white')
        buf = io.BytesIO()
        img.save(buf, format='PNG')
        buf.seek(0)
        return base64.b64encode(buf.read()).decode()
    except Exception as e:
        print('[Apollo] qrcode falhou:', e)
        return ''


@app.route('/api/qrcode/gerar', methods=['POST'])
def gerar_qr_endpoint():
    if not usuario_logado():
        return resp_erro('Não autenticado.', 401)
    d = request.get_json() or {}
    b64 = gerar_qr_base64(
        d.get('codigo', ''), d.get('nome', ''), d.get('categoria', ''),
        d.get('tipo', 'etiqueta'), d.get('quantidade', 0),
        d.get('estoque_atual', 0), d.get('estoque_minimo', 0),
        d.get('data', datetime.now().strftime('%Y-%m-%d')),
        d.get('fornecedor', ''), d.get('fornecedor_email', ''), d.get('fornecedor_telefone', '')
    )
    if not b64:
        return resp_erro('Nao foi possivel gerar o QR Code. Verifique: pip install qrcode[pil]')
    return resp_ok({'qr_base64': b64})





# ═══════════════════════════════════════════════════════════
#  GERAÇÃO DE EXCEL (openpyxl — código da equipe)
# ═══════════════════════════════════════════════════════════
@app.route('/api/relatorio/excel', methods=['GET'])
def gerar_excel():
    if not usuario_logado() or not somente_admin():
        return resp_erro('Acesso negado. Somente administradores.', 403)

    data_de  = request.args.get('de')
    data_ate = request.args.get('ate')
    sheets   = request.args.get('sheets', 'movimentacoes,estoque').split(',')

    wb = Workbook()
    wb.remove(wb.active)   # remove aba default

    # ── Estilos (baseado no planilhaExcell.py da equipe) ──
    cor_cabecalho = PatternFill('solid', fgColor='0A1F44')   # azul navy Apollo
    cor_linha_par = PatternFill('solid', fgColor='E8EEF8')
    cor_verde      = PatternFill('solid', fgColor='DCFCE7')
    cor_vermelho   = PatternFill('solid', fgColor='FEE2E2')
    cor_amarelo    = PatternFill('solid', fgColor='FEF9C3')

    fonte_header = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    fonte_dados  = Font(name='Arial', size=10, color='1E293B')
    fonte_titulo = Font(name='Arial', size=14, bold=True, color='0A1F44')

    alinha_centro = Alignment(horizontal='center', vertical='center')
    alinha_esq    = Alignment(horizontal='left',   vertical='center')

    borda = Border(
        left  =Side(style='thin', color='CBD5E1'),
        right =Side(style='thin', color='CBD5E1'),
        top   =Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    def estilizar_header(ws, colunas, linha=1):
        for col_idx, titulo in enumerate(colunas, 1):
            cell = ws.cell(row=linha, column=col_idx, value=titulo)
            cell.font      = fonte_header
            cell.fill      = cor_cabecalho
            cell.alignment = alinha_centro
            cell.border    = borda
        ws.row_dimensions[linha].height = 28

    def estilizar_celula(cell, cor_fill=None, centro=False):
        cell.font      = fonte_dados
        cell.alignment = alinha_centro if centro else alinha_esq
        cell.border    = borda
        if cor_fill:
            cell.fill = cor_fill

    conn = get_db()

    # ── ABA: MOVIMENTAÇÕES ──────────────────────────────────
    if 'movimentacoes' in sheets:
        ws = wb.create_sheet('Movimentações')

        # Título
        ws.merge_cells('A1:H1')
        t = ws['A1']
        t.value     = f'Apollo Logística — Movimentações  |  Gerado em {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        t.font      = fonte_titulo
        t.alignment = alinha_centro
        ws.row_dimensions[1].height = 32

        colunas = ['Código','Produto','Categoria','Tipo','Quantidade','Estoque Após','Data','Operador']
        estilizar_header(ws, colunas, linha=2)

        query = 'SELECT * FROM movimentacoes WHERE 1=1'
        params = []
        if data_de:  query += ' AND data >= ?'; params.append(data_de)
        if data_ate: query += ' AND data <= ?'; params.append(data_ate)
        query += ' ORDER BY id DESC'
        movs = conn.execute(query, params).fetchall()

        for i, m in enumerate(movs):
            row_num  = i + 3
            par      = (i % 2 == 0)
            fill_row = cor_linha_par if par else None
            tipo_fill= cor_verde if m['tipo'] == 'entrada' else cor_vermelho

            dados = [
                m['codigo_produto'], m['nome_produto'], m['categoria'],
                'Entrada ▲' if m['tipo'] == 'entrada' else 'Saída ▼',
                m['quantidade'], m['estoque_apos'],
                datetime.strptime(m['data'], '%Y-%m-%d').strftime('%d/%m/%Y'),
                m['operador'] or '—'
            ]
            for col_idx, val in enumerate(dados, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                if col_idx == 4:   # coluna Tipo
                    estilizar_celula(cell, tipo_fill, centro=True)
                elif col_idx in (5, 6):
                    estilizar_celula(cell, fill_row, centro=True)
                else:
                    estilizar_celula(cell, fill_row)

        larguras = [12, 28, 18, 12, 12, 14, 14, 16]
        for i, w in enumerate(larguras, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = 'A3'

    # ── ABA: ESTOQUE ATUAL ──────────────────────────────────
    if 'estoque' in sheets:
        ws = wb.create_sheet('Estoque Atual')

        ws.merge_cells('A1:H1')
        t = ws['A1']
        t.value     = f'Apollo Logística — Estoque Atual  |  {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        t.font      = fonte_titulo
        t.alignment = alinha_centro
        ws.row_dimensions[1].height = 32

        colunas = ['Código','Produto','Categoria','Estoque Atual','Estoque Mínimo','Status','Fornecedor','E-mail Fornecedor','Telefone Fornecedor']
        estilizar_header(ws, colunas, linha=2)

        prods = conn.execute('SELECT * FROM estoque ORDER BY nome').fetchall()
        for i, p in enumerate(prods):
            row_num  = i + 3
            par      = (i % 2 == 0)
            fill_row = cor_linha_par if par else None
            baixo    = p['quantidade_atual'] <= p['estoque_minimo']
            fill_st  = cor_vermelho if p['quantidade_atual'] == 0 else (cor_amarelo if baixo else cor_verde)

            dados = [
                p['codigo'], p['nome'], p['categoria'],
                p['quantidade_atual'], p['estoque_minimo'],
                'ZERADO 🔴' if p['quantidade_atual'] == 0 else ('BAIXO ⚠️' if baixo else 'OK ✅'),
                p['fornecedor'] or '—', p['fornecedor_email'] or '—',
                p['fornecedor_telefone'] or '—'
            ]
            for col_idx, val in enumerate(dados, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                if col_idx == 6:
                    estilizar_celula(cell, fill_st, centro=True)
                elif col_idx in (4, 5):
                    estilizar_celula(cell, fill_row, centro=True)
                else:
                    estilizar_celula(cell, fill_row)

        larguras = [12, 28, 18, 14, 16, 12, 24, 30, 20]
        for i, w in enumerate(larguras, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = 'A3'

    # ── ABA: ALERTAS ────────────────────────────────────────
    if 'alertas' in sheets:
        ws = wb.create_sheet('Alertas Estoque')

        ws.merge_cells('A1:G1')
        t = ws['A1']
        t.value     = f'Apollo Logística — Produtos Críticos  |  {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        t.font      = fonte_titulo
        t.alignment = alinha_centro
        ws.row_dimensions[1].height = 32

        colunas = ['Código','Produto','Categoria','Estoque Atual','Estoque Mínimo','Situação','Fornecedor']
        estilizar_header(ws, colunas, linha=2)

        baixos = conn.execute(
            'SELECT * FROM estoque WHERE quantidade_atual <= estoque_minimo ORDER BY quantidade_atual'
        ).fetchall()

        if not baixos:
            cell = ws.cell(row=3, column=1, value='Nenhum produto em estoque crítico ✅')
            cell.font = Font(name='Arial', size=11, color='15803D', bold=True)
            ws.merge_cells('A3:G3')
        else:
            for i, p in enumerate(baixos):
                row_num = i + 3
                fill_st = cor_vermelho if p['quantidade_atual'] == 0 else cor_amarelo
                dados = [
                    p['codigo'], p['nome'], p['categoria'],
                    p['quantidade_atual'], p['estoque_minimo'],
                    'ZERADO 🔴' if p['quantidade_atual'] == 0 else 'CRÍTICO 🟡',
                    p['fornecedor'] or '—'
                ]
                for col_idx, val in enumerate(dados, 1):
                    cell = ws.cell(row=row_num, column=col_idx, value=val)
                    estilizar_celula(cell, fill_st if col_idx == 6 else None, centro=(col_idx in (4,5,6)))

        larguras = [12, 28, 18, 14, 16, 14, 24]
        for i, w in enumerate(larguras, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ── ABA: USUÁRIOS ────────────────────────────────────────
    if 'usuarios' in sheets:
        ws = wb.create_sheet('Usuários')

        ws.merge_cells('A1:D1')
        t = ws['A1']
        t.value     = f'Apollo Logística — Usuários  |  {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        t.font      = fonte_titulo
        t.alignment = alinha_centro
        ws.row_dimensions[1].height = 32

        colunas = ['Usuário','E-mail','Perfil','Data de Cadastro']
        estilizar_header(ws, colunas, linha=2)

        users = conn.execute(
            'SELECT usuario, email, role, criado_em FROM usuarios ORDER BY criado_em DESC'
        ).fetchall()

        cor_admin = PatternFill('solid', fgColor='EDE9FE')
        cor_func  = PatternFill('solid', fgColor='E0EAFF')

        for i, u in enumerate(users):
            row_num = i + 3
            fill    = cor_admin if u['role'] == 'admin' else cor_func
            dados   = [
                u['usuario'], u['email'],
                'Administrador' if u['role'] == 'admin' else 'Funcionário',
                datetime.fromisoformat(u['criado_em']).strftime('%d/%m/%Y %H:%M')
            ]
            for col_idx, val in enumerate(dados, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                estilizar_celula(cell, fill if col_idx == 3 else None)

        for i, w in enumerate([20, 32, 16, 20], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    conn.close()

    # ── Salva e retorna o arquivo ────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from flask import send_file
    nome_arq = f"Apollo_Relatorio_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=nome_arq
    )


# ═══════════════════════════════════════════════════════════
#  DASHBOARD — RESUMO
# ═══════════════════════════════════════════════════════════
@app.route('/api/resumo', methods=['GET'])
def resumo():
    if not usuario_logado():
        return resp_erro('Não autenticado.', 401)
    hoje = datetime.now().strftime('%Y-%m-%d')
    conn = get_db()
    c    = conn.cursor()

    total_produtos = c.execute('SELECT COUNT(*) FROM estoque').fetchone()[0]
    entradas_hoje  = c.execute(
        "SELECT COALESCE(SUM(quantidade),0) FROM movimentacoes WHERE tipo='entrada' AND data=?", (hoje,)
    ).fetchone()[0]
    saidas_hoje    = c.execute(
        "SELECT COALESCE(SUM(quantidade),0) FROM movimentacoes WHERE tipo='saida' AND data=?", (hoje,)
    ).fetchone()[0]
    estoque_baixo  = c.execute(
        'SELECT COUNT(*) FROM estoque WHERE quantidade_atual <= estoque_minimo'
    ).fetchone()[0]

    conn.close()
    return resp_ok({
        'total_produtos': total_produtos,
        'entradas_hoje':  entradas_hoje,
        'saidas_hoje':    saidas_hoje,
        'estoque_baixo':  estoque_baixo
    })


# ═══════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════

# Garante que o banco é criado ao importar o módulo (dev e gunicorn)
with app.app_context():
    init_db()

if __name__ == '__main__':
    print("=" * 50)
    print("  🚀 Apollo Logística — Servidor iniciado!")
    print("  Acesse: http://localhost:5000")
    print("=" * 50)
    app.run(debug=False, host='0.0.0.0', port=5000)
